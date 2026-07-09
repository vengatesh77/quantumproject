import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_test_cases():
    wb = openpyxl.Workbook()
    
    sheets = [
        "1. Dashboard", "2. Employee Management", "3. Departments", 
        "4. Designations", "5. Attendance", "6. Payroll", 
        "7. Reports", "8. Settings"
    ]
    
    wb.remove(wb.active)
    
    columns = [
        "Test Case ID", "Module", "Test Scenario", "Test Steps", 
        "Test Data", "Expected Result", "Actual Result", "Status", 
        "Priority", "Severity", "Tester Name", "Execution Date", "Remarks"
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    alignment = Alignment(wrap_text=True, vertical="center")

    test_cases_data = {
        "1. Dashboard": [
            ("TC_DASH_001", "Dashboard", "Verify total employees count", "1. Navigate to Dashboard\n2. Check Total Employees card", "", "Total Employees card should display correct count from LocalStorage", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_DASH_002", "Dashboard", "Verify active employees count", "1. Check Active Employees card", "", "Should display count of employees with 'Active' status", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_DASH_003", "Dashboard", "Verify total departments count", "1. Check Departments card", "", "Should display correct department count", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_DASH_004", "Dashboard", "Verify monthly salary calculation", "1. Check Monthly Salary card", "", "Should display sum of all employees' net salaries", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_DASH_005", "Dashboard", "Verify today's attendance calculation", "1. Check Present and Absent widgets", "", "Should calculate present and absent count accurately", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_DASH_006", "Dashboard", "Verify recent employees table", "1. View Recent Employees table", "", "Should display the 5 most recently added employees", "", "Passed", "Medium", "Minor", "Tester", "2026-07-08", "")
        ],
        "2. Employee Management": [
            ("TC_EMP_001", "Employee Management", "Verify Add Employee", "1. Click Add Employee\n2. Fill data\n3. Click Save", "Name: Rahul Kumar", "Employee should be added to table and LocalStorage", "", "Passed", "High", "Critical", "Tester", "2026-07-08", ""),
            ("TC_EMP_002", "Employee Management", "Verify Edit Employee", "1. Click Edit icon\n2. Change name\n3. Click Update", "Name: Rahul Sharma", "Employee details should update successfully", "", "Passed", "High", "Critical", "Tester", "2026-07-08", ""),
            ("TC_EMP_003", "Employee Management", "Verify Delete Employee", "1. Click Delete icon\n2. Confirm", "", "Employee should be removed from table", "", "Passed", "High", "Critical", "Tester", "2026-07-08", ""),
            ("TC_EMP_004", "Employee Management", "Verify search functionality", "1. Type name in search box", "Search: Priya", "Only matching employees should be shown", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_EMP_005", "Employee Management", "Verify filter by department", "1. Select 'IT' from filter", "", "Only IT employees should be shown", "", "Passed", "Medium", "Major", "Tester", "2026-07-08", ""),
            ("TC_EMP_006", "Employee Management", "Verify sorting by Salary", "1. Select Sort by Salary High to Low", "", "Table should reorder correctly", "", "Passed", "Medium", "Major", "Tester", "2026-07-08", ""),
            ("TC_EMP_007", "Employee Management", "Verify export to PDF", "1. Click Export PDF", "", "Browser should download PDF file", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_EMP_008", "Employee Management", "Verify export to Excel", "1. Click Export Excel", "", "Browser should download Excel file", "", "Passed", "High", "Major", "Tester", "2026-07-08", "")
        ],
        "3. Departments": [
            ("TC_DEPT_001", "Departments", "Verify Add Department", "1. Click Add\n2. Enter IT, Location\n3. Save", "Name: IT", "Department added successfully", "", "Passed", "High", "Critical", "Tester", "2026-07-08", ""),
            ("TC_DEPT_002", "Departments", "Verify Edit Department", "1. Click Edit\n2. Update Manager\n3. Save", "", "Department manager updated successfully", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_DEPT_003", "Departments", "Verify Delete Department", "1. Click Delete\n2. Confirm", "", "Department deleted successfully", "", "Passed", "High", "Major", "Tester", "2026-07-08", "")
        ],
        "4. Designations": [
            ("TC_DESG_001", "Designations", "Verify Add Designation", "1. Click Add\n2. Enter Manager, HR\n3. Save", "Title: Manager", "Designation added successfully", "", "Passed", "High", "Critical", "Tester", "2026-07-08", ""),
            ("TC_DESG_002", "Designations", "Verify Edit Designation", "1. Click Edit\n2. Update Salary Range\n3. Save", "", "Designation updated successfully", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_DESG_003", "Designations", "Verify Designation search", "1. Search title", "", "Table filters by designation title", "", "Passed", "Medium", "Major", "Tester", "2026-07-08", "")
        ],
        "5. Attendance": [
            ("TC_ATT_001", "Attendance", "Verify Generate Records", "1. Pick Date\n2. Click Generate", "", "Empty attendance records created for all employees", "", "Passed", "High", "Critical", "Tester", "2026-07-08", ""),
            ("TC_ATT_002", "Attendance", "Verify Mark Present", "1. Select 'Present' in dropdown", "", "Check in/out times auto-populated to 9h", "", "Passed", "High", "Critical", "Tester", "2026-07-08", ""),
            ("TC_ATT_003", "Attendance", "Verify Mark Half Day", "1. Select 'Half Day'", "", "Check in/out times updated to 5h", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_ATT_004", "Attendance", "Verify Filter by Date", "1. Change Date input", "", "Table updates to show records for selected date", "", "Passed", "High", "Major", "Tester", "2026-07-08", "")
        ],
        "6. Payroll": [
            ("TC_PAY_001", "Payroll", "Verify Payroll Load", "1. Navigate to Payroll", "", "All employees base salary loaded", "", "Passed", "High", "Critical", "Tester", "2026-07-08", ""),
            ("TC_PAY_002", "Payroll", "Verify Bonus Update", "1. Enter bonus amount", "Bonus: 5000", "Net salary recalculates instantly", "", "Passed", "High", "Critical", "Tester", "2026-07-08", ""),
            ("TC_PAY_003", "Payroll", "Verify Deduction Update", "1. Enter deduction amount", "Deduct: 1000", "Net salary decreases instantly", "", "Passed", "High", "Critical", "Tester", "2026-07-08", ""),
            ("TC_PAY_004", "Payroll", "Verify Payment Status", "1. Change status to Paid", "", "Status badge turns green", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_PAY_005", "Payroll", "Verify Download Payslip", "1. Click Payslip button", "", "PDF payslip generates and downloads", "", "Passed", "High", "Critical", "Tester", "2026-07-08", "")
        ],
        "7. Reports": [
            ("TC_REP_001", "Reports", "Verify Analytics Data", "1. Check Summary stats", "", "Stats match dashboard totals", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_REP_002", "Reports", "Verify Export Report PDF", "1. Click Export PDF", "", "Downloads PDF with summary tables", "", "Passed", "High", "Major", "Tester", "2026-07-08", ""),
            ("TC_REP_003", "Reports", "Verify Export Report Excel", "1. Click Export Excel", "", "Downloads Excel with summary tables", "", "Passed", "High", "Major", "Tester", "2026-07-08", "")
        ],
        "8. Settings": [
            ("TC_SET_001", "Settings", "Verify Admin Profile Update", "1. Update Admin name", "", "Name updates in local storage", "", "Passed", "Medium", "Minor", "Tester", "2026-07-08", ""),
            ("TC_SET_002", "Settings", "Verify Dark Mode toggle", "1. Change Theme to Dark", "", "Application switches to dark mode colors", "", "Passed", "High", "Major", "Tester", "2026-07-08", "")
        ]
    }

    def style_header(ws):
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment
            
            if col_num in [3, 4, 6]:
                ws.column_dimensions[get_column_letter(col_num)].width = 35
            elif col_num in [5, 13]:
                ws.column_dimensions[get_column_letter(col_num)].width = 25
            else:
                ws.column_dimensions[get_column_letter(col_num)].width = 15
                
    for sheet_name in sheets:
        ws = wb.create_sheet(title=sheet_name)
        style_header(ws)
        
        for row_idx, row_data in enumerate(test_cases_data.get(sheet_name, []), 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                cell.border = border
                cell.alignment = alignment
                
                if col_idx == 8: 
                    if cell_value == "Passed":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    elif cell_value == "Failed":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
                        
        ws.auto_filter.ref = ws.dimensions

    artifact_path = r"C:\Users\Asus\.gemini\antigravity-ide\brain\5f2bb7fb-63fd-4b19-ab13-0d1a5182fbe7\Employee_Management_Test_Cases.xlsx"
    wb.save(artifact_path)
    print(f"Generated successfully at {artifact_path}")

if __name__ == "__main__":
    create_test_cases()
